# ==============================================================================
# Table T1: subtype changes in budget, GHG abatement, and biodiversity
#           contribution under the highest carbon-price and biodiversity-price
#           scenarios.
#
# Inputs:
#   - Figure 01 raw workbook: area difference relative to the zero-price run.
#   - Figure 03 raw workbook: GHG and biodiversity differences relative to the
#     zero-price run.
#   - Figure 04 raw workbook: budget difference relative to the zero-price run.
#
# Output:
#   - A publication-oriented Excel workbook in paper4/tables.
# ==============================================================================

import os
import sys
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from tools.price_slice_utils import DATA_DIR, OUT_DIR, format_thousands


YEAR = 2025
AREA_PATH = DATA_DIR / f"01_Area_Delta_vs_Zero_raw_data_{YEAR}.xlsx"
CONTRIBUTION_PATH = DATA_DIR / f"03_Contribution_Delta_vs_Zero_raw_data_{YEAR}.xlsx"
BUDGET_PATH = DATA_DIR / f"04_Budget_Delta_vs_Zero_raw_data_{YEAR}.xlsx"

TABLE_DIR = OUT_DIR.parent / "tables"
OUT_PATH = TABLE_DIR / f"T1_Subtype_Budget_GHG_Bio_{YEAR}.xlsx"

TABLE_TITLE = (
    "Table T1. Changes in budget, GHG abatement, biodiversity contribution, "
    "and area by subtype under the highest carbon-price and biodiversity-price "
    "scenarios."
)
TABLE_CAPTION = (
    f"Values are changes relative to the zero-price run for {YEAR}. "
    "The carbon-price drive uses the maximum carbon price "
    "with biodiversity price set to zero, and the biodiversity-price drive uses "
    "the maximum biodiversity price with carbon price set to zero. Budget follows "
    "Figure 04 dvar-based accounting and is reported in billion AUD yr^-1. "
    "GHG abatement and biodiversity contribution follow Figure 03 and are reported "
    "in Mt CO2e yr^-1 and Mha yr^-1, respectively. Area follows Figure 01 and is "
    "reported in Mha. "
    "Transition GHG is reported as a separate row under agricultural land-use."
)

GHG_METRIC = "GHGAbatementChange_vs_ZeroPrice_MtCO2e"
BIO_METRIC = "BiodiversityContributionChange_vs_ZeroPrice_MhaYr"

BUDGET_COL = "Budget (billion AUD yr^-1)"
AREA_COL = "Area (Mha)"
GHG_COL = "GHG abatement (Mt CO2e yr^-1)"
BIO_COL = "Biodiversity contribution (Mha yr^-1)"

DOMAIN_LABELS = {
    "Agricultural land-use": "Agricultural land-use",
    "Ag management": "Agricultural management",
    "Non-ag": "Non-agricultural land-use",
}

DOMAIN_ORDER = {
    "Agricultural land-use": 1,
    "Ag management": 2,
    "Non-ag": 3,
}

SUBTYPE_ORDER_BY_AREA = {
    "Agricultural land-use": [
        "Unallocated - natural land",
        "Unallocated - modified land",
        "Crops",
        "Modified livestock",
        "Natural Livestock",
        "Transition",
    ],
    "Ag management": [
        "Methane reduction (livestock)",
        "Agricultural technology (fertiliser)",
        "Early dry-season savanna burning",
        "Agricultural technology (energy)",
        "Biochar (soil amendment)",
        "Managed regeneration (beef)",
        "Managed regeneration (sheep)",
    ],
    "Non-ag": [
        "Environmental plantings (mixed species)",
        "Riparian buffer restoration (mixed species)",
        "Agroforestry (mixed species + sheep)",
        "Agroforestry (mixed species + beef)",
        "Carbon plantings (monoculture)",
        "Farm forestry (hardwood timber + sheep)",
        "Farm forestry (hardwood timber + beef)",
        "BECCS (Bioenergy with Carbon Capture and Storage)",
        "Destocked - natural land",
    ],
}


def subtype_order(area_type, category):
    base_order = SUBTYPE_ORDER_BY_AREA.get(area_type, [])
    if category in base_order:
        return base_order.index(category) + 1
    return len(base_order) + 1


def require_input(path, producer):
    if not path.is_file():
        raise FileNotFoundError(
            f"Required workbook not found: {path}\n"
            f"Run {producer} first, then re-run this table script."
        )


def load_inputs():
    require_input(AREA_PATH, "01_Area_Delta_vs_Zero.py")
    require_input(CONTRIBUTION_PATH, "03_Contribution_Delta_vs_Zero.py")
    require_input(BUDGET_PATH, "04_Budget_Delta_vs_Zero.py")

    area = pd.read_excel(AREA_PATH, sheet_name="AreaLong")
    contribution = pd.read_excel(CONTRIBUTION_PATH, sheet_name="ContributionLong")
    budget = pd.read_excel(BUDGET_PATH, sheet_name="NetEconLong")
    return area, contribution, budget


def selected_scenarios(contribution):
    max_cp = float(contribution.loc[contribution["PriceType"].eq("CarbonPrice"), "Price"].max())
    max_bp = float(contribution.loc[contribution["PriceType"].eq("BioPrice"), "Price"].max())

    return pd.DataFrame(
        [
            {
                "Scenario": "Carbon price drive",
                "PriceType": "CarbonPrice",
                "Price": max_cp,
                "Carbon price (AU$/tCO2e yr^-1)": max_cp,
                "Biodiversity price (AU$/ha yr^-1)": 0.0,
                "Price setting": f"Carbon price = {format_thousands(max_cp)}; biodiversity price = 0",
                "Scenario order": 1,
            },
            {
                "Scenario": "Biodiversity price drive",
                "PriceType": "BioPrice",
                "Price": max_bp,
                "Carbon price (AU$/tCO2e yr^-1)": 0.0,
                "Biodiversity price (AU$/ha yr^-1)": max_bp,
                "Price setting": f"Carbon price = 0; biodiversity price = {format_thousands(max_bp)}",
                "Scenario order": 2,
            },
        ]
    )


def make_metric_table(area, contribution, budget, scenarios):
    scenario_keys = scenarios[["PriceType", "Price"]]

    contribution = contribution.merge(scenario_keys, on=["PriceType", "Price"], how="inner")
    contribution = contribution[contribution["MetricType"].isin([GHG_METRIC, BIO_METRIC])].copy()
    contribution_wide = contribution.pivot_table(
        index=["PriceType", "Price", "AreaType", "Category"],
        columns="MetricType",
        values="ContributionValue",
        aggfunc="sum",
        fill_value=0.0,
    ).reset_index()
    contribution_wide = contribution_wide.rename(
        columns={
            GHG_METRIC: GHG_COL,
            BIO_METRIC: BIO_COL,
        }
    )

    budget = budget.merge(scenario_keys, on=["PriceType", "Price"], how="inner")
    budget = budget[
        [
            "PriceType",
            "Price",
            "AreaType",
            "Category",
            "NetEconChange_vs_ZeroPrice_BAUD",
        ]
    ].copy()
    budget = budget.rename(columns={"NetEconChange_vs_ZeroPrice_BAUD": BUDGET_COL})

    area = area.merge(scenario_keys, on=["PriceType", "Price"], how="inner")
    area = area[
        [
            "PriceType",
            "Price",
            "AreaType",
            "Category",
            "AreaChange_vs_ZeroPrice_Mha",
        ]
    ].copy()
    area = area.rename(columns={"AreaChange_vs_ZeroPrice_Mha": AREA_COL})

    metrics = contribution_wide.merge(
        budget,
        on=["PriceType", "Price", "AreaType", "Category"],
        how="outer",
    ).merge(
        area,
        on=["PriceType", "Price", "AreaType", "Category"],
        how="outer",
    )

    for col in [AREA_COL, BUDGET_COL, GHG_COL, BIO_COL]:
        if col not in metrics.columns:
            metrics[col] = 0.0
        metrics[col] = metrics[col].fillna(0.0)

    subtype_keys = metrics[["AreaType", "Category"]].drop_duplicates()
    complete = scenarios.merge(subtype_keys, how="cross")
    complete = complete.merge(
        metrics,
        on=["PriceType", "Price", "AreaType", "Category"],
        how="left",
    )

    for col in [AREA_COL, BUDGET_COL, GHG_COL, BIO_COL]:
        complete[col] = complete[col].fillna(0.0)
        complete.loc[np.isclose(complete[col], 0.0, atol=1e-6), col] = 0.0

    complete["Domain"] = complete["AreaType"].map(DOMAIN_LABELS).fillna(complete["AreaType"])
    complete["Subtype"] = complete["Category"]
    complete["Domain order"] = complete["AreaType"].map(DOMAIN_ORDER).fillna(99)
    complete["Subtype order"] = complete.apply(
        lambda row: subtype_order(row["AreaType"], row["Category"]),
        axis=1,
    )

    keep = [
        "Scenario",
        "Price setting",
        "Carbon price (AU$/tCO2e yr^-1)",
        "Biodiversity price (AU$/ha yr^-1)",
        "Domain",
        "Subtype",
        AREA_COL,
        BUDGET_COL,
        GHG_COL,
        BIO_COL,
        "Scenario order",
        "Domain order",
        "Subtype order",
    ]
    complete = complete[keep].copy()
    complete = complete.sort_values(["Scenario order", "Domain order", "Subtype order", "Subtype"])
    return complete.reset_index(drop=True)


def make_summary_table(table):
    scenario_cols = [
        "Scenario order",
        "Scenario",
        "Price setting",
        "Carbon price (AU$/tCO2e yr^-1)",
        "Biodiversity price (AU$/ha yr^-1)",
    ]
    group_cols = [
        "Scenario order",
        "Domain order",
        "Scenario",
        "Price setting",
        "Carbon price (AU$/tCO2e yr^-1)",
        "Biodiversity price (AU$/ha yr^-1)",
        "Domain",
    ]
    metric_cols = [AREA_COL, BUDGET_COL, GHG_COL, BIO_COL]
    summary = table.groupby(group_cols, as_index=False)[metric_cols].sum()
    totals = table.groupby(scenario_cols, as_index=False)[metric_cols].sum()
    totals["Domain order"] = 99
    totals["Domain"] = "Total"
    summary = pd.concat([summary, totals], ignore_index=True)
    summary = summary.sort_values(["Scenario order", "Domain order"]).reset_index(drop=True)
    return summary.drop(columns=["Scenario order", "Domain order"])


def write_workbook(table, summary, scenarios):
    TABLE_DIR.mkdir(parents=True, exist_ok=True)

    table_for_excel = table.drop(columns=["Scenario order", "Domain order", "Subtype order"])

    metadata = pd.DataFrame(
        [
            ["Table title", TABLE_TITLE],
            ["Caption", TABLE_CAPTION],
            ["Year", YEAR],
            ["Carbon-price scenario", scenarios.loc[0, "Price setting"]],
            ["Biodiversity-price scenario", scenarios.loc[1, "Price setting"]],
            ["Budget source", str(BUDGET_PATH)],
            ["Contribution source", str(CONTRIBUTION_PATH)],
            ["Area source", str(AREA_PATH)],
            ["Budget accounting", "Figure 04 dvar-based budget difference relative to the zero-price run."],
            ["Contribution accounting", "Figure 03 GHG and biodiversity differences relative to the zero-price run."],
            ["Area accounting", "Figure 01 area difference relative to the zero-price run."],
            ["GHG sign convention", "Positive values indicate GHG abatement."],
        ],
        columns=["Field", "Value"],
    )

    with pd.ExcelWriter(OUT_PATH, engine="openpyxl") as writer:
        pd.DataFrame({"A": [TABLE_TITLE, TABLE_CAPTION, ""]}).to_excel(
            writer,
            sheet_name="T1_Table",
            index=False,
            header=False,
            startrow=0,
        )
        table_for_excel.to_excel(writer, sheet_name="T1_Table", index=False, startrow=4)
        summary.to_excel(writer, sheet_name="T1_Summary", index=False)
        metadata.to_excel(writer, sheet_name="T1_Metadata", index=False)

    style_workbook(OUT_PATH, table_for_excel)
    print(f"Saved: {OUT_PATH}")


def style_workbook(path, table_for_excel):
    wb = load_workbook(path)
    ws = wb["T1_Table"]

    last_col = len(table_for_excel.columns)
    last_row = 5 + len(table_for_excel)

    title_font = Font(name="Arial", size=12, bold=True, color="000000")
    body_font = Font(name="Arial", size=10, color="000000")
    header_font = Font(name="Arial", size=10, bold=True, color="000000")
    header_fill = PatternFill("solid", fgColor="D9EAD3")
    thin_gray = Side(style="thin", color="B7B7B7")
    border = Border(top=thin_gray, bottom=thin_gray, left=thin_gray, right=thin_gray)

    add_domain_sheets(wb, table_for_excel, body_font, header_font, header_fill, border)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    ws["A1"].font = title_font
    ws["A2"].font = body_font
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[2].height = 58

    for row in ws.iter_rows(min_row=5, max_row=last_row, max_col=last_col):
        for cell in row:
            cell.font = body_font
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    for cell in ws[5]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    metric_headers = {AREA_COL, BUDGET_COL, GHG_COL, BIO_COL}
    price_headers = {
        "Carbon price (AU$/tCO2e yr^-1)",
        "Biodiversity price (AU$/ha yr^-1)",
    }
    header_by_col = {cell.column: cell.value for cell in ws[5]}
    for col_idx, header in header_by_col.items():
        if header in metric_headers:
            for row in range(6, last_row + 1):
                ws.cell(row=row, column=col_idx).number_format = '#,##0.000'
        elif header in price_headers:
            for row in range(6, last_row + 1):
                ws.cell(row=row, column=col_idx).number_format = '#,##0'

    widths = {
        "A": 24,
        "B": 42,
        "C": 16,
        "D": 18,
        "E": 26,
        "F": 48,
        "G": 18,
        "H": 18,
        "I": 22,
        "J": 22,
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = "A6"
    ws.sheet_view.showGridLines = False

    table_ref = f"A5:{get_column_letter(last_col)}{last_row}"
    excel_table = Table(displayName="T1_SubtypeMetrics", ref=table_ref)
    excel_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(excel_table)

    for sheet_name in ["T1_Summary", "T1_Metadata"]:
        sheet = wb[sheet_name]
        sheet.sheet_view.showGridLines = False
        for row in sheet.iter_rows():
            for cell in row:
                cell.font = body_font
                cell.alignment = Alignment(vertical="center", wrap_text=True)
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
        for col_idx in range(1, sheet.max_column + 1):
            sheet.column_dimensions[get_column_letter(col_idx)].width = 24
            header = sheet.cell(row=1, column=col_idx).value
            if header in metric_headers:
                for row_idx in range(2, sheet.max_row + 1):
                    sheet.cell(row=row_idx, column=col_idx).number_format = '#,##0.000'
            elif header in price_headers:
                for row_idx in range(2, sheet.max_row + 1):
                    sheet.cell(row=row_idx, column=col_idx).number_format = '#,##0'
        if sheet_name == "T1_Metadata":
            sheet.column_dimensions["B"].width = 110
        sheet.freeze_panes = "A2"

    wb.save(path)


def add_domain_sheets(wb, table_for_excel, body_font, header_font, header_fill, border):
    domain_sheets = [
        ("ag", "Agricultural land-use"),
        ("agmgt", "Agricultural management"),
        ("non-ag", "Non-agricultural land-use"),
    ]
    headers = [
        "Type",
        AREA_COL,
        BUDGET_COL,
        GHG_COL,
        BIO_COL,
    ]

    for sheet_name, domain in domain_sheets:
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(sheet_name)
        ws.sheet_view.showGridLines = False

        row_idx = 1
        domain_df = table_for_excel[table_for_excel["Domain"].eq(domain)].copy()
        for scenario in pd.unique(table_for_excel["Scenario"]):
            scenario_df = domain_df[domain_df["Scenario"].eq(scenario)].copy()
            if scenario_df.empty:
                continue

            scenario_df = scenario_df.sort_values(BUDGET_COL, ascending=False)
            price_setting = scenario_df["Price setting"].iloc[0]
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=5)
            title_cell = ws.cell(row=row_idx, column=1)
            title_cell.value = f"{scenario}: {price_setting}"
            title_cell.font = header_font
            title_cell.alignment = Alignment(vertical="center", wrap_text=True)
            row_idx += 1

            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            row_idx += 1

            for _, source_row in scenario_df.iterrows():
                values = [
                    source_row["Subtype"],
                    source_row[AREA_COL],
                    source_row[BUDGET_COL],
                    source_row[GHG_COL],
                    source_row[BIO_COL],
                ]
                for col_idx, value in enumerate(values, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    cell.font = body_font
                    cell.border = border
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                    if col_idx > 1:
                        cell.number_format = '#,##0.000'
                row_idx += 1

            row_idx += 2

        ws.column_dimensions["A"].width = 50
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 22
        ws.column_dimensions["D"].width = 24
        ws.column_dimensions["E"].width = 28
        ws.freeze_panes = "A3"


def main():
    area, contribution, budget = load_inputs()
    scenarios = selected_scenarios(contribution)
    table = make_metric_table(area, contribution, budget, scenarios)
    summary = make_summary_table(table)
    write_workbook(table, summary, scenarios)

    print("\nSelected scenarios:")
    for _, row in scenarios.iterrows():
        print(f"  {row['Scenario']}: {row['Price setting']}")


if __name__ == "__main__":
    main()
