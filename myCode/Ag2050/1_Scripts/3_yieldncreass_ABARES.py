
from tools import predict_growth_index
import pandas as pd
import matplotlib.pyplot as plt
import math
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
import copy
import matplotlib.collections as mcoll

# ---------- 配置 ----------
def preprocessed_df(df):
    col0 = df.columns[0]
    col1 = df.columns[1]

    # 1) 提取 '-' 前面的年份并转为整数
    df['Year'] = (
        df[col0]
        .astype(str)  # 确保为字符串
        .str.strip()  # 去掉前后空白
        .str.split('-', n=1)  # 按第一个 '-' 拆分
        .str[0]  # 取左侧部分
    )

    # 尝试把 Year 转成整数，无法转换的会变成 NaN
    df['Year'] = pd.to_numeric(df['Year'], errors='coerce').astype('Int64')
    # 2) 重命名并把第二列转为数值型
    df['Cost'] = pd.to_numeric(df[col1], errors='coerce')
    df.drop(df.columns[:2], axis=1, inplace=True)
    return df


# ---------- 主程序 ----------

sheet_configs = [
    ('table_c2', 'Crop Productivity'),
    ('table_18', 'Dairy Productivity'),
    ('table_c4', 'Sheep Productivity'),
    ('table_c5', 'Beef Productivity')
]

all_plots = []
all_results = []
output_file = "../0_original_data/yieldincreases_ABARES.xlsx"
for sheet_name, var_name in sheet_configs:
    df = pd.read_excel("../0_original_data/ABARES_productivity.xlsx",
                       sheet_name=sheet_name, usecols="A,B", skiprows=1)
    df_processed = preprocessed_df(df)
    # df_processed = df_processed.query("Year >= 2000")
    ax, df_result = predict_growth_index(df_processed, var_name=var_name,base_year=1988,model='ETS')

    all_plots.append((var_name, ax))
    df_result['Variable'] = var_name
    all_results.append(df_result)

# 合并预测结果
combined_df_result = pd.concat(all_results, ignore_index=False)

# 设置每行2个图
n_plots = len(all_plots)
ncols = 2
nrows = math.ceil(n_plots / ncols)

fig, axs = plt.subplots(nrows, ncols, figsize=(12, 4 * nrows), constrained_layout=True)
axs = axs.flatten()  # 转为一维，方便索引

for i, (ax_target, (title, ax_source)) in enumerate(zip(axs, all_plots)):

    # ✅ 方法1：复制阴影带（更稳健的方式）
    for collection in ax_source.collections:
        # 检查是否是 PolyCollection (fill_between 创建的)
        if collection.__class__.__name__ == 'PolyCollection':
            # 获取路径
            paths = collection.get_paths()
            if len(paths) > 0 and len(paths[0].vertices) > 0:
                vertices = paths[0].vertices

                # fill_between 的顶点结构：[x下, y下] -> [x上逆序, y上逆序]
                n = len(vertices)

                # 找分界点：通常是第一个 x 值开始重复的地方
                x_vals = vertices[:, 0]

                # 简单分割：前半部分是下边界，后半部分（逆序）是上边界
                split_idx = n // 2

                # 尝试更精确的分割点（找到 x 开始递减的位置）
                for idx in range(1, n):
                    if x_vals[idx] < x_vals[idx - 1]:
                        split_idx = idx
                        break

                x_lower = vertices[:split_idx, 0]
                y_lower = vertices[:split_idx, 1]
                x_upper = vertices[split_idx:, 0][::-1]  # 逆序
                y_upper = vertices[split_idx:, 1][::-1]

                # 确保长度一致
                min_len = min(len(x_lower), len(x_upper))
                if min_len > 0:
                    ax_target.fill_between(
                        x_lower[:min_len],
                        y_lower[:min_len],
                        y_upper[:min_len],
                        color=collection.get_facecolor()[0],
                        alpha=collection.get_alpha() if collection.get_alpha() else 0.2,
                        label=collection.get_label() if collection.get_label() else None,
                        zorder=collection.get_zorder()
                    )

        # 复制散点（PathCollection，如 scatter 创建的）
        elif collection.__class__.__name__ == 'PathCollection':
            offsets = collection.get_offsets()
            if offsets.size > 0:
                x, y = offsets[:, 0], offsets[:, 1]
                facecolors = collection.get_facecolor()
                sizes = collection.get_sizes()

                ax_target.scatter(
                    x, y,
                    label=collection.get_label() if collection.get_label() else None,
                    color=facecolors[0] if len(facecolors) > 0 else 'black',
                    s=sizes[0] if len(sizes) > 0 else 20,
                    zorder=collection.get_zorder()
                )

    # 复制线条
    for line in ax_source.get_lines():
        ax_target.plot(
            *line.get_data(),
            label=line.get_label() if line.get_label() and not line.get_label().startswith('_') else None,
            linestyle=line.get_linestyle(),
            color=line.get_color(),
            linewidth=line.get_linewidth(),
            zorder=line.get_zorder(),
            alpha=line.get_alpha() if line.get_alpha() else 1.0
        )

    # 设置标题、标签、范围
    ax_target.set_title(title, fontsize=14, weight='bold')
    ax_target.set_xlabel(ax_source.get_xlabel(), fontsize=12)
    ax_target.set_ylabel(ax_source.get_ylabel(), fontsize=12)
    ax_target.set_xlim(ax_source.get_xlim())
    ax_target.set_ylim(ax_source.get_ylim())
    ax_target.grid(True, alpha=0.3)

    # 复制水平/垂直参考线
    for line in ax_source.lines:
        data_x, data_y = line.get_data()
        # 检测水平线（y值恒定）
        if len(set(data_y)) == 1:
            ax_target.axhline(
                data_y[0],
                linestyle=line.get_linestyle(),
                color=line.get_color(),
                linewidth=line.get_linewidth(),
                alpha=line.get_alpha() if line.get_alpha() else 1.0,
                zorder=line.get_zorder()
            )

    # ✅ 只有第一个子图显示图例
    if i == 0:
        # 去重图例（避免重复标签）
        handles, labels = ax_target.get_legend_handles_labels()
        by_label = dict(zip(labels, handles))
        ax_target.legend(
            by_label.values(),
            by_label.keys(),
            loc='upper left',
            frameon=True,
            fontsize=9
        )

# 隐藏多余的子图
for j in range(i + 1, len(axs)):
    axs[j].axis('off')


plt.suptitle("Productivity Predictions", fontsize=16)
plt.savefig("productivity_plots.png", dpi=300)
plt.show()

# 保存预测结果
# combined_df_result.to_excel("combined_productivity_results.xlsx", index=False)


# 配置输出文件名
out_path = Path("FLC_multipliers_by_scenario.xlsx")

# 确保年份为索引且为整型（或可被识别的索引）
df_result = combined_df_result
df_result.index = df_result.index.astype(int)

# 获取模板的列（保留 MultiIndex 结构，如果有）
df_template = pd.read_excel(output_file, header=[0, 1], sheet_name=1)
template_cols = df_template.columns

var_to_columns = {
    'Crop Productivity': [
        'Apples', 'Citrus', 'Cotton', 'Grapes', 'Hay', 'Nuts', 'Other non-cereal crops',
        'Pears', 'Plantation fruit', 'Rice', 'Stone fruit', 'Sugar',
        'Summer cereals', 'Summer legumes', 'Summer oilseeds',
        'Tropical stone fruit', 'Vegetables',
        'Winter cereals', 'Winter legumes', 'Winter oilseeds'
    ],
    'Dairy Productivity': [
        'Dairy - modified land', 'Dairy - natural land'
    ],
    'Beef Productivity': [
        'BEEF - MODIFIED LAND LEXP', 'BEEF - NATURAL LAND LEXP',
        'BEEF - MODIFIED LAND MEAT', 'BEEF - NATURAL LAND MEAT'
    ],
    'Sheep Productivity': [
        'SHEEP - MODIFIED LAND LEXP', 'SHEEP - MODIFIED LAND MEAT', 'SHEEP - MODIFIED LAND WOOL',
        'SHEEP - NATURAL LAND LEXP', 'SHEEP - NATURAL LAND MEAT', 'SHEEP - NATURAL LAND WOOL'
    ]
}

def expand_to_multicolumn(cols):
    return [('dry', c.upper()) for c in cols] + [('irr', c.upper()) for c in cols]

# 修正后的映射表（适用于 MultiIndex）
var_to_columns = {
    k: expand_to_multicolumn(v)
    for k, v in var_to_columns.items()
}

# 年份顺序（写入时使用 df_result 的索引）
years = sorted(df_result.index.unique())
required_scenarios = ['Low', 'Medium', 'High', 'Very_High']
sheets_created = []

p = Path(output_file)
mode = 'a' if p.exists() else 'w'
with pd.ExcelWriter(output_file, engine="openpyxl", mode=mode) as writer:
    for scenario in required_scenarios:
        print(f"处理场景: {scenario}")

        df_sheet = pd.DataFrame(index=years, columns=template_cols, dtype=float)
        df_sheet.index.name = 'Year'

        for y in years:
            for var_name, col_list in var_to_columns.items():
                value_row = df_result.query("Variable == @var_name and Year == @y")
                if not value_row.empty:
                    val = value_row[scenario].values[0]
                    for col in col_list:
                        if col in df_sheet.columns:
                            df_sheet.loc[y, col] = val
                        else:
                            print(f"[跳过] 列 {col} 不在模板中")
        df_sheet.fillna(1, inplace=True)
        df_out = df_sheet.reset_index()
        df_out.columns.names = [None, None]

        # 写入 Excel，创建新 sheet
        wb = writer.book
        ws = wb.create_sheet(title=scenario.lower())
        sheets_created.append(scenario.lower())

        col_level_0 = ['Year'] + [col[0] for col in df_out.columns[1:]]
        col_level_1 = [''] + [col[1] for col in df_out.columns[1:]]
        ws.append(col_level_0)
        ws.append(col_level_1)

        for row in df_out.itertuples(index=False):
            ws.append(row)

    # 安全删除默认 Sheet
    if 'Sheet' in writer.book.sheetnames and len(sheets_created) > 0:
        del writer.book['Sheet']

print(f"已保存至 {output_file}")