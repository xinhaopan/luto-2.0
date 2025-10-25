
from tools import predit_growth_index
import pandas as pd
import matplotlib.pyplot as plt
import math
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook

# ---------- 配置 ----------
def preprocessed_df(df):
    col0 = df.columns[0]
    col1 = df.columns[1]

    # 1) 提取 '-' 前面的年份并转为整数
    df['Year'] = (
        df[col0]
        .astype(str)  # 确保为字符串
        .str.strip()  # 去掉前后空白
        .str.split('-', n=1)  # 按第一个 '-' 拆分
        .str[0]  # 取左侧部分
    )

    # 尝试把 Year 转成整数，无法转换的会变成 NaN
    df['Year'] = pd.to_numeric(df['Year'], errors='coerce').astype('Int64')
    # 2) 重命名并把第二列转为数值型
    df['Cost'] = pd.to_numeric(df[col1], errors='coerce')
    df.drop(df.columns[:2], axis=1, inplace=True)
    return df


# ---------- 主程序 ----------
output_file = "yieldncreases_ABARES.xlsx"
sheet_configs = [
    ('table_c2', 'Crop Productivity'),
    ('table_18', 'Dairy Productivity'),
    ('table_c4', 'Sheep Productivity'),
    ('table_c5', 'Beef Productivity')
]

all_plots = []
all_results = []

for sheet_name, var_name in sheet_configs:
    df = pd.read_excel('../0_original_data/ABARES_productivity.xlsx',
                       sheet_name=sheet_name, usecols="A,B", skiprows=1)
    df_processed = preprocessed_df(df)
    ax, df_result = predit_growth_index(df_processed, var_name=var_name)

    all_plots.append((var_name, ax))
    df_result['Variable'] = var_name
    all_results.append(df_result)

# 合并预测结果
combined_df_result = pd.concat(all_results, ignore_index=False)

# 设置每行2个图
n_plots = len(all_plots)
ncols = 2
nrows = math.ceil(n_plots / ncols)

fig, axs = plt.subplots(nrows, ncols, figsize=(12, 4 * nrows), constrained_layout=True)
axs = axs.flatten()  # 转为一维，方便索引

for i, (ax_target, (title, ax_source)) in enumerate(zip(axs, all_plots)):
    # 复制线
    for line in ax_source.get_lines():
        ax_target.plot(*line.get_data(), label=line.get_label(),
                       linestyle=line.get_linestyle(), color=line.get_color())

    # 复制点
    for col in ax_source.collections:
        offsets = col.get_offsets()
        if offsets.size == 0:
            continue
        x, y = offsets[:, 0], offsets[:, 1]
        ax_target.scatter(x, y, label=col.get_label(),
                          color=col.get_facecolor()[0],
                          s=col.get_sizes()[0] if col.get_sizes().size > 0 else 20)

    # 设置标题、标签、x范围
    ax_target.set_title(title)
    ax_target.set_xlabel(ax_source.get_xlabel())
    ax_target.set_ylabel(ax_source.get_ylabel())
    ax_target.set_xlim(df['Year'][0], 2050)

    # ✅ 只有第一个子图显示图例
    if i == 0:
        ax_target.legend()

# 多余的 subplot 去掉
for i in range(len(all_plots), len(axs)):
    fig.delaxes(axs[i])

plt.suptitle("Productivity Predictions", fontsize=16)
plt.savefig("combined_productivity_plots.png", dpi=300)
plt.show()

# 保存预测结果
combined_df_result.to_excel("combined_productivity_results.xlsx", index=False)


# 配置输出文件名
out_path = Path("FLC_multipliers_by_scenario.xlsx")

# 确保年份为索引且为整型（或可被识别的索引）
df_result = combined_df_result
df_result.index = df_result.index.astype(int)

# 获取模板的列（保留 MultiIndex 结构，如果有）
df_template = pd.read_csv('../0_original_data/yieldincreases_bau2022.csv', header=[0, 1])
template_cols = df_template.columns

var_to_columns = {
    'Crop Productivity': [
        'Apples', 'Citrus', 'Cotton', 'Grapes', 'Hay', 'Nuts', 'Other non-cereal crops',
        'Pears', 'Plantation fruit', 'Rice', 'Stone fruit', 'Sugar',
        'Summer cereals', 'Summer legumes', 'Summer oilseeds',
        'Tropical stone fruit', 'Vegetables',
        'Winter cereals', 'Winter legumes', 'Winter oilseeds'
    ],
    'Dairy Productivity': [
        'Dairy - modified land', 'Dairy - natural land'
    ],
    'Beef Productivity': [
        'BEEF - MODIFIED LAND LEXP', 'BEEF - NATURAL LAND LEXP',
        'BEEF - MODIFIED LAND MEAT', 'BEEF - NATURAL LAND MEAT'
    ],
    'Sheep Productivity': [
        'SHEEP - MODIFIED LAND LEXP', 'SHEEP - MODIFIED LAND MEAT', 'SHEEP - MODIFIED LAND WOOL',
        'SHEEP - NATURAL LAND LEXP', 'SHEEP - NATURAL LAND MEAT', 'SHEEP - NATURAL LAND WOOL'
    ]
}

def expand_to_multicolumn(cols):
    return [('dry', c.upper()) for c in cols] + [('irr', c.upper()) for c in cols]

# 修正后的映射表（适用于 MultiIndex）
var_to_columns = {
    k: expand_to_multicolumn(v)
    for k, v in var_to_columns.items()
}

# 年份顺序（写入时使用 df_result 的索引）
years = sorted(df_result.index.unique())
required_scenarios = ['Low', 'Medium', 'High', 'Very_High']
sheets_created = []

with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
    for scenario in required_scenarios:
        print(f"处理场景: {scenario}")

        df_sheet = pd.DataFrame(index=years, columns=template_cols, dtype=float)
        df_sheet.index.name = 'Year'

        for y in years:
            for var_name, col_list in var_to_columns.items():
                value_row = df_result.query("Variable == @var_name and Year == @y")
                if not value_row.empty:
                    val = value_row[scenario].values[0]
                    for col in col_list:
                        if col in df_sheet.columns:
                            df_sheet.loc[y, col] = val
                        else:
                            print(f"[跳过] 列 {col} 不在模板中")
        df_sheet.fillna(1, inplace=True)
        df_out = df_sheet.reset_index()
        df_out.columns.names = [None, None]

        # 写入 Excel，创建新 sheet
        wb = writer.book
        ws = wb.create_sheet(title=scenario.lower())
        sheets_created.append(scenario.lower())

        col_level_0 = ['Year'] + [col[0] for col in df_out.columns[1:]]
        col_level_1 = [''] + [col[1] for col in df_out.columns[1:]]
        ws.append(col_level_0)
        ws.append(col_level_1)

        for row in df_out.itertuples(index=False):
            ws.append(row)

    # 安全删除默认 Sheet
    if 'Sheet' in writer.book.sheetnames and len(sheets_created) > 0:
        del writer.book['Sheet']

print(f"已保存至 {output_file}")